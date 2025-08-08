import os
import dbase
import cb
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

pd.set_option("display.max_rows", None)
pd.set_option("display.max_columns", None)
pd.set_option("display.width", 0)
pd.set_option("display.max_colwidth", None)

startswith = "AEG"

def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

def format_nominal_columns(ws, nominal_cols):
    header = {cell.value: cell.column for cell in ws[1]}
    for col_name in nominal_cols:
        if col_name in header:
            col_idx = header[col_name]
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = '#,##0'

def point_1():
    print("=== GOLD POINT 1 ===")

    def kol_defcb():
        df_cb = cb.CB()
        df_cb = df_cb[df_cb["NAMA_SLS"].str.startswith((startswith))].reset_index(drop=True)
        return df_cb[["PMA", "KD_SLS", "NAMA_SLS"]]

    def hitung_target(df_cb):
        file_path = os.path.join(os.getcwd(), "R09 PRIANGAN TIMUR.xlsm")
        try:
            df_target = pd.read_excel(file_path, sheet_name="SALESMAN LIST", usecols=["KD_SLS", "TARGET"])
        except Exception as e:
            print(f"❌ Gagal membaca file TARGET: {e}")
            df_cb["TARGET"] = 0
            return df_cb

        df_merged = df_cb.merge(df_target, on="KD_SLS", how="left")
        df_merged["TARGET"] = df_merged["TARGET"].fillna(0)
        return df_merged.drop_duplicates()

    def omset(df_merged):
        dfksni = dbase.KSNI()
        dfksni = dfksni[dfksni["NAMA SLS2"].str.startswith((startswith))].reset_index(drop=True)
        df_grouped = dfksni.groupby("KD SLS2")[["VALUE"]].sum().reset_index()
        df_grouped = df_grouped.rename(columns={"KD SLS2": "KD_SLS"})
        df_grouped["VALUE"] = df_grouped["VALUE"].fillna(0)
        return df_merged.merge(df_grouped, on="KD_SLS", how="left")

    def hitung_insentif(df_result):
        df = df_result.copy()
        df["TARGET"] = df["TARGET"].fillna(0).astype(float)
        df["VALUE"] = df["VALUE"].fillna(0).astype(float)

        def kalkulasi_insentif(row):
            if row["TARGET"] > 0 and row["VALUE"] >= row["TARGET"]:
                return 1_500_000
            elif row["TARGET"] > 0 and row["VALUE"] >= row["TARGET"] * 0.975:
                return 1_000_000
            elif row["TARGET"] > 0 and row["VALUE"] >= row["TARGET"] * 0.95:
                return 750_000
            else:
                return 0

        df["INSENTIF"] = df.apply(kalkulasi_insentif, axis=1)

        # print(df)  # Cetak hasil akhir (penuh karena sudah set_option)
        return df

    # === Eksekusi ===
    df_cb = kol_defcb()
    df_merged = hitung_target(df_cb)
    df_final = omset(df_merged)
    df_point1 = hitung_insentif(df_final)

    output_folder = os.path.join(os.getcwd(), "ScoreCard")
    os.makedirs(output_folder, exist_ok=True)

    for pma, group in df_point1.groupby("PMA"):
        nama_file = f"{pma}.xlsx"
        file_path = os.path.join(output_folder, nama_file)
        sheet_name = "GOLD"
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(title=sheet_name)
            for col_idx, col_name in enumerate(group.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            for row_idx, row in enumerate(group.values, 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        else:
            group.to_excel(file_path, sheet_name=sheet_name, index=False)
            wb = load_workbook(file_path)
            ws = wb[sheet_name]
        auto_adjust_column_width(ws)
        format_nominal_columns(ws, nominal_cols=["TARGET", "VALUE", "INSENTIF"])

        wb.save(file_path)
        print(f"✅ File disimpan: {file_path}")

    return df_point1

def point_2():
    print("=== GOLD POINT 2 ===")

        # === Fungsi untuk mengambil kolom default ===
    def kol_defcb():
        df_cb = cb.CB()
        df_cb = df_cb[df_cb["NAMA_SLS"].str.startswith(startswith)]
        df_cb = df_cb[["PMA", "KD_SLS", "NAMA_SLS"]]
        df_cb = df_cb.drop_duplicates(subset=["PMA", "KD_SLS", "NAMA_SLS"]).reset_index(drop=True)
        return df_cb
        
    def CSD2K(df_cb):
        def ambil_target():
            file_path = os.path.join(os.getcwd(), "R09 PRIANGAN TIMUR.xlsm")
            try:
                df_target = pd.read_excel(
                    file_path,
                    sheet_name="SALESMAN LIST",
                    usecols=["KD_SLS", "CSD-E02K"]
                )
                df_target = df_target.rename(columns={"CSD-E02K": "TVALUE"})
                df_target["TVALUE"] = df_target["TVALUE"].fillna(0).astype(int)
                return df_target
            except Exception as e:
                print(f"❌ Gagal membaca file TARGET: {e}")
                return pd.DataFrame(columns=["KD_SLS", "TVALUE"])
        dfcsd2k = dbase.KSNI()
        dfcsd2k = dfcsd2k[dfcsd2k["NAMA SLS2"].str.startswith(startswith)].reset_index(drop=True)
        fdfcsd2k = [309944, 309945, 309999, 307821, 307822]
        dfcsd2k = dfcsd2k[dfcsd2k["KD_BRG"].isin(fdfcsd2k)]
        dfcsd2k["CSD2K"] = 1
        df_sum = dfcsd2k.groupby("KD SLS2")[["CSD2K", "VALUE"]].sum().reset_index()
        df_sum = df_sum.rename(columns={"KD SLS2": "KD_SLS"})
        df_merged = df_cb.merge(df_sum, on="KD_SLS", how="left")
        df_merged["CSD2K"] = df_merged["CSD2K"].fillna(0).astype(int)
        df_merged["VALUE"] = df_merged["VALUE"].fillna(0).astype(int)
        df_merged["TCSD2K"] = 50
        df_target = ambil_target()
        df_merged = df_merged.merge(df_target, on="KD_SLS", how="left")
        df_merged["TVALUE"] = df_merged["TVALUE"].fillna(0).astype(int)
        df_merged["SCSD2K"] = ((df_merged["CSD2K"] >= df_merged["TCSD2K"]) &
                            (df_merged["VALUE"] >= df_merged["TVALUE"])).astype(int)
        df_merged["VALUE"] = df_merged["VALUE"].apply(lambda x: f"{x:,}")
        df_merged["TVALUE"] = df_merged["TVALUE"].apply(lambda x: f"{x:,}")
        kol_tambahan = ["TCSD2K", "CSD2K", "TVALUE", "VALUE", "SCSD2K"]
        return df_merged[kol_tambahan]

    def NXT2K(df_cb):
        def ambil_target():
            file_path = os.path.join(os.getcwd(), "R09 PRIANGAN TIMUR.xlsm")
            try:
                df_target = pd.read_excel(
                    file_path,
                    sheet_name="SALESMAN LIST",
                    usecols=["KD_SLS", "NXT-B02K"]
                )
                df_target = df_target.rename(columns={"NXT-B02K": "TVALUE"})
                df_target["TVALUE"] = df_target["TVALUE"].fillna(0).astype(int)
                return df_target
            except Exception as e:
                print(f"❌ Gagal membaca file TARGET: {e}")
                return pd.DataFrame(columns=["KD_SLS", "TVALUE"])
        dfNXT2K = dbase.KSNI()
        dfNXT2K = dfNXT2K[dfNXT2K["NAMA SLS2"].str.startswith(startswith)].reset_index(drop=True)
        fdfNXT2K = [303174]
        dfNXT2K = dfNXT2K[dfNXT2K["KD_BRG"].isin(fdfNXT2K)]
        dfNXT2K["NXT2K"] = 1
        df_sum = dfNXT2K.groupby("KD SLS2")[["NXT2K", "VALUE"]].sum().reset_index()
        df_sum = df_sum.rename(columns={"KD SLS2": "KD_SLS"})
        df_merged = df_cb.merge(df_sum, on="KD_SLS", how="left")
        df_merged["NXT2K"] = df_merged["NXT2K"].fillna(0).astype(int)
        df_merged["VALUE"] = df_merged["VALUE"].fillna(0).astype(int)
        df_merged["TNXT2K"] = 50
        df_target = ambil_target()
        df_merged = df_merged.merge(df_target, on="KD_SLS", how="left")
        df_merged["TVALUE"] = df_merged["TVALUE"].fillna(0).astype(int)
        df_merged["SNXT2K"] = ((df_merged["NXT2K"] >= df_merged["TNXT2K"]) &
                            (df_merged["VALUE"] >= df_merged["TVALUE"])).astype(int)
        df_merged["VALUE"] = df_merged["VALUE"].apply(lambda x: f"{x:,}")
        df_merged["TVALUE"] = df_merged["TVALUE"].apply(lambda x: f"{x:,}")
        kol_tambahan = ["TNXT2K", "NXT2K", "TVALUE", "VALUE", "SNXT2K"]
        return df_merged[kol_tambahan]
    
    def TB2K(df_cb):
        def ambil_target():
            file_path = os.path.join(os.getcwd(), "R09 PRIANGAN TIMUR.xlsm")
            try:
                df_target = pd.read_excel(
                    file_path,
                    sheet_name="SALESMAN LIST",
                    usecols=["KD_SLS", "TBK-E02K"]
                )
                df_target = df_target.rename(columns={"TBK-E02K": "TVALUE"})
                df_target["TVALUE"] = df_target["TVALUE"].fillna(0).astype(int)
                return df_target
            except Exception as e:
                print(f"❌ Gagal membaca file TARGET: {e}")
                return pd.DataFrame(columns=["KD_SLS", "TVALUE"])
        dfTB2K = dbase.KSNI()
        dfTB2K = dfTB2K[dfTB2K["NAMA SLS2"].str.startswith(startswith)].reset_index(drop=True)
        fdftb2k = [300196, 300812, 301924, 301938]
        dfTB2K = dfTB2K[dfTB2K["KD_BRG"].isin(fdftb2k)]
        dfTB2K["TB2K"] = 1
        df_sum = dfTB2K.groupby("KD SLS2")[["TB2K", "VALUE"]].sum().reset_index()
        df_sum = df_sum.rename(columns={"KD SLS2": "KD_SLS"})
        df_merged = df_cb.merge(df_sum, on="KD_SLS", how="left")
        df_merged["TB2K"] = df_merged["TB2K"].fillna(0).astype(int)
        df_merged["VALUE"] = df_merged["VALUE"].fillna(0).astype(int)
        df_merged["TTB2K"] = 50
        df_target = ambil_target()
        df_merged = df_merged.merge(df_target, on="KD_SLS", how="left")
        df_merged["TVALUE"] = df_merged["TVALUE"].fillna(0).astype(int)
        df_merged["STB2K"] = ((df_merged["TB2K"] >= df_merged["TTB2K"]) &
                            (df_merged["VALUE"] >= df_merged["TVALUE"])).astype(int)
        df_merged["VALUE"] = df_merged["VALUE"].apply(lambda x: f"{x:,}")
        df_merged["TVALUE"] = df_merged["TVALUE"].apply(lambda x: f"{x:,}")
        kol_tambahan = ["TTB2K", "TB2K", "TVALUE", "VALUE", "STB2K"]
        return df_merged[kol_tambahan]
    
    def NXB2K(df_cb):
        def ambil_target():
            file_path = os.path.join(os.getcwd(), "R09 PRIANGAN TIMUR.xlsm")
            try:
                df_target = pd.read_excel(
                    file_path,
                    sheet_name="SALESMAN LIST",
                    usecols=["KD_SLS", "NXC-E02K"]
                )
                df_target = df_target.rename(columns={"NXC-E02K": "TVALUE"})
                df_target["TVALUE"] = df_target["TVALUE"].fillna(0).astype(int)
                return df_target
            except Exception as e:
                print(f"❌ Gagal membaca file TARGET: {e}")
                return pd.DataFrame(columns=["KD_SLS", "TVALUE"])
        dfNXB2K = dbase.KSNI()
        dfNXB2K = dfNXB2K[dfNXB2K["NAMA SLS2"].str.startswith(startswith)].reset_index(drop=True)
        fdfnxb2k = [303176]
        dfNXB2K = dfNXB2K[dfNXB2K["KD_BRG"].isin(fdfnxb2k)]
        dfNXB2K["NXB2K"] = 1
        df_sum = dfNXB2K.groupby("KD SLS2")[["NXB2K", "VALUE"]].sum().reset_index()
        df_sum = df_sum.rename(columns={"KD SLS2": "KD_SLS"})
        df_merged = df_cb.merge(df_sum, on="KD_SLS", how="left")
        df_merged["NXB2K"] = df_merged["NXB2K"].fillna(0).astype(int)
        df_merged["VALUE"] = df_merged["VALUE"].fillna(0).astype(int)
        df_merged["TNXB2K"] = 50
        df_target = ambil_target()
        df_merged = df_merged.merge(df_target, on="KD_SLS", how="left")
        df_merged["TVALUE"] = df_merged["TVALUE"].fillna(0).astype(int)
        df_merged["SNXB2K"] = ((df_merged["NXB2K"] >= df_merged["TNXB2K"]) &
                            (df_merged["VALUE"] >= df_merged["TVALUE"])).astype(int)
        df_merged["VALUE"] = df_merged["VALUE"].apply(lambda x: f"{x:,}")
        df_merged["TVALUE"] = df_merged["TVALUE"].apply(lambda x: f"{x:,}")
        kol_tambahan = ["TNXB2K", "NXB2K", "TVALUE", "VALUE", "SNXB2K"]
        return df_merged[kol_tambahan]

    def reward_platp2(df_result):
        df_result["TFocuS"] = (
            df_result["SCSD2K"] +
            df_result["SNXT2K"] +
            df_result["STB2K"] +
            df_result["SNXB2K"]
        )

        def hitung_insentif(focus):
            if focus == 2:
                return 1_000_000
            elif focus == 3:
                return 1_750_000
            elif focus >= 4:
                return 2_000_000
            else:
                return 0

        df_result["Insentive"] = df_result["TFocuS"].apply(hitung_insentif)
        return df_result

    # === Jalankan seluruh proses ===
    df_cb = kol_defcb()
    df_csd2k = CSD2K(df_cb)
    df_nxt2k = NXT2K(df_cb)
    df_tb2k = TB2K(df_cb)
    df_nxb2k = NXB2K(df_cb)

    df_result = pd.concat([df_cb, df_csd2k, df_nxt2k, df_tb2k, df_nxb2k], axis=1)
    df_result = reward_platp2(df_result)

    print(df_result)
