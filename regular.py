import os
import dbase
import cb
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

# Konfigurasi tampilan pandas
pd.set_option("display.max_rows", None)
pd.set_option("display.max_columns", None)
pd.set_option("display.width", 0)
pd.set_option("display.max_colwidth", None)

# === Fungsi bantu: atur lebar kolom secara otomatis ===
def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

def format_nominal_columns(ws, nominal_cols):
    header = {cell.value: cell.column for cell in ws[1]}
    for col_name in nominal_cols:
        if col_name in header:
            col_idx = header[col_name]
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = '#,##0'

# === Fungsi utama REGULAR POINT 1 ===
def point_1():
    print("=== REGULAR POINT 1 ===")

    def kol_defcb():
        df_cb = cb.CB()
        df_cb = df_cb[df_cb["NAMA_SLS"].str.startswith(("TX2D1", "TX2D2", "TX2DA"))].reset_index(drop=True)
        return df_cb[["PMA", "KD_SLS", "NAMA_SLS"]]

    def hitung_target(df_cb):
        file_path = os.path.join(os.getcwd(), "R09 PRIANGAN TIMUR.xlsm")
        try:
            df_target = pd.read_excel(file_path, sheet_name="SALESMAN LIST", usecols=["KD_SLS", "TARGET"])
        except Exception as e:
            print(f"❌ Gagal membaca file TARGET: {e}")
            df_cb["TARGET"] = 0
            return df_cb

        df_merged = df_cb.merge(df_target, on="KD_SLS", how="left")
        df_merged["TARGET"] = df_merged["TARGET"].fillna(0)
        return df_merged.drop_duplicates()

    def omset(df_merged):
        dfksni = dbase.KSNI()
        dfksni = dfksni[dfksni["NAMA SLS2"].str.startswith(("TX2D1", "TX2D2", "TX2DA"))].reset_index(drop=True)
        df_grouped = dfksni.groupby("KD SLS2")[["VALUE"]].sum().reset_index()
        df_grouped["VALUE"] = df_grouped["VALUE"].fillna(0)
        df_grouped = df_grouped.rename(columns={"KD SLS2": "KD_SLS"})
        return df_merged.merge(df_grouped, on="KD_SLS", how="left")

    def hitung_insentif(df_result):
        df = df_result.copy()
        df["TARGET"] = pd.to_numeric(df["TARGET"], errors="coerce").fillna(0)
        df["VALUE"] = pd.to_numeric(df["VALUE"], errors="coerce").fillna(0)

        def apply_minimal_target(row):
            nama = row["NAMA_SLS"]
            if nama.startswith(("TX2D1", "TX2D2")) and row["TARGET"] < 40_000_000:
                return 40_000_000
            elif nama.startswith("TX2DA") and row["TARGET"] < 60_000_000:
                return 60_000_000
            return row["TARGET"]

        df["TARGET"] = df.apply(apply_minimal_target, axis=1)

        def kalkulasi_insentif(row):
            if row["TARGET"] > 0 and row["VALUE"] >= row["TARGET"]:
                return 500_000
            elif row["TARGET"] > 0 and row["VALUE"] >= row["TARGET"] * 0.975:
                return 350_000
            elif row["TARGET"] > 0 and row["VALUE"] >= row["TARGET"] * 0.95:
                return 250_000
            else:
                return 0

        df["INSENTIF"] = df.apply(kalkulasi_insentif, axis=1)

        # print(df)  # debug print hasil akhir
        return df

    # === Eksekusi pipeline ===
    df_cb = kol_defcb()
    df_merged = hitung_target(df_cb)
    df_final = omset(df_merged)
    df_point1 = hitung_insentif(df_final)

    # === Simpan ke Excel per PMA ===
    output_folder = os.path.join(os.getcwd(), "ScoreCard")
    os.makedirs(output_folder, exist_ok=True)

    for pma, group in df_point1.groupby("PMA"):
        nama_file = f"{pma}.xlsx"
        file_path = os.path.join(output_folder, nama_file)
        sheet_name = "REGULAR"
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(title=sheet_name)
            for col_idx, col_name in enumerate(group.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            for row_idx, row in enumerate(group.values, 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        else:
            group.to_excel(file_path, sheet_name=sheet_name, index=False)
            wb = load_workbook(file_path)
            ws = wb[sheet_name]
        auto_adjust_column_width(ws)
        format_nominal_columns(ws, nominal_cols=["TARGET", "VALUE", "INSENTIF"])

        wb.save(file_path)
        print(f"✅ File disimpan: {file_path}")

    return df_point1

